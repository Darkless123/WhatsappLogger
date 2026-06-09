import os
import json
from datetime import datetime
from openpyxl import load_workbook
from pydantic import BaseModel, Field
from openai import OpenAI

# Define the data structure for validation
class WhatsAppIssueExtractor(BaseModel):
    client_name: str = Field(description="The company or person who sent the message.")
    system: str = Field(description="The system or module impacted (e.g., eAppTax, CFF, Portal, DB).")
    clean_summary: str = Field(description="1-sentence clean technical description of the issue.")
    apparent_cause: str = Field(description="Engineering guess at the root cause (e.g., SP Timeout, reCAPTCHA fail).")
    severity: str = Field(description="CRITICAL, WARNING, or INFO.")

def append_to_manday_log(file_path: str, data: WhatsAppIssueExtractor):
    if not os.path.exists(file_path):
        print(f"Error: Could not find workbook at {file_path}")
        return

    wb = load_workbook(file_path)
    target_tab = "Mandays Utilization_New"
    
    if target_tab in wb.sheetnames:
        ws = wb[target_tab]
    else:
        ws = wb.create_sheet(title=target_tab)
        ws.append(["Log Timestamp", "Client / Source", "Impacted System", "Technical Summary", "Likely Cause", "Severity", "Status"])

    new_row = [
        datetime.now().strftime("%Y-%m-%d %H:%M"),
        data.client_name,
        data.system.upper(),
        data.clean_summary,
        data.apparent_cause,
        data.severity.upper(),
        "OPEN"
    ]
    
    ws.append(new_row)
    wb.save(file_path)
    print(f"✓ Successfully logged via DeepSeek to '{target_tab}' tab!")

# --- Main DeepSeek Integration ---
def parse_whatsapp_with_deepseek(raw_whatsapp_text: str, excel_path: str):
    # Point the standard OpenAI client to DeepSeek's API endpoint
    client = OpenAI(
        base_url="https://api.deepseek.com",
        api_key=os.environ.get("DEEPSEEK_API_KEY") 
    )
    
    # For DeepSeek JSON mode, we clearly define the structural requirements in the system prompt
    system_instructions = (
        "You are an expert technical triage assistant. Parse unstructured WhatsApp chats from clients, "
        "extract the crucial engineering details, and return them strictly as a valid JSON object.\n\n"
        "The output JSON must strictly contain these exact keys:\n"
        "- 'client_name'\n"
        "- 'system'\n"
        "- 'clean_summary'\n"
        "- 'apparent_cause'\n"
        "- 'severity'\n"
        "Do not include any explanation or markdown formatting."
    )
    
    # Call DeepSeek with json_object format enabled
    response = client.chat.completions.create(
        model="deepseek-v4-flash", # Fast and highly optimal for text extraction/parsing
        messages=[
            {"role": "system", "content": system_instructions},
            {"role": "user", "content": f"Parse this message:\n\n{raw_whatsapp_text}"}
        ],
        response_format={"type": "json_object"},
        temperature=0.1
    )
    
    # Extract raw content string and deserialize it
    raw_json_string = response.choices[0].message.content
    parsed_json = json.loads(raw_json_string)
    
    # Push into Pydantic model to ensure data structural health before writing to Excel
    validated_data = WhatsAppIssueExtractor(**parsed_json)
    
    append_to_manday_log(excel_path, validated_data)

if __name__ == "__main__":
    whatsapp_chat = """
    [11:15 AM] Tan: Kevin, can check the CFF production environment? 
    Our console app keeps hitting a connection timeout when trying to auto-remove the public files. 
    It ran fine yesterday but today it just hangs indefinitely on the cleanup thread.
    """
    
    # Target your local sheet
    workbook_path = "IssueLog.xlsx"
    parse_whatsapp_with_deepseek(whatsapp_chat, workbook_path)