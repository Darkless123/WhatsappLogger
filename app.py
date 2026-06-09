import os
import json
import traceback
import copy
from datetime import datetime
from flask import Flask, request, jsonify
from flask_cors import CORS
from openpyxl import load_workbook
from pydantic import BaseModel, Field
from openai import OpenAI
from openpyxl.utils import range_boundaries, get_column_letter

app = Flask(__name__)
CORS(app)

# 1. Strict Schema Blueprint matching your columns
class MandayLogExtractor(BaseModel):
    reported_by: str = Field(default="WhatsApp Client", description="Name of the person/client reporting the issue.")
    system: str = Field(default="SMART", description="System impacted: Must be exactly 'SMART', 'CFF', or 'eApp & eTax'.")
    category: str = Field(default="Issue Support", description="Classification: Must be exactly 'Issue Support', 'Clarification', or 'Enhancement'.")
    priority: int = Field(default=2, description="Priority level: 1 for critical system crashes/500 errors, 2 for standard operational bugs.")
    details: str = Field(description="A clean, concise 1-sentence engineering summary of the problem.")
    internal_remark: str = Field(default="", description="Technical deduction, hypothesis, or data clues for internal notes.")
        
def append_to_actual_manday_sheet(file_path: str, data: MandayLogExtractor, raw_timestamp: str):
    if not os.path.exists(file_path):
        return f"Error: Workbook file not found at {file_path}"

    try:
        wb = load_workbook(file_path)
        target_tab = "SMART&CFF"
        
        if target_tab not in wb.sheetnames:
            return f"Error: Tab '{target_tab}' not found in workbook."
            
        ws = wb[target_tab]

        insert_row = None
        current_max_no = 0
        start_row = 4

        # Dynamic scanner to find your TOTAL MAN-DAYS anchor block row
        for r in range(start_row, ws.max_row + 15):
            val_a = ws.cell(row=r, column=1).value
            val_b = ws.cell(row=r, column=2).value

            if isinstance(val_a, (int, float)):
                current_max_no = max(current_max_no, int(val_a))

            if val_b and "TOTAL MAN-DAYS" in str(val_b).upper():
                insert_row = r
                break

        # Fallback safety default
        if not insert_row:
            insert_row = ws.max_row

        # Squeeze a clean row into the spreadsheet data region
        ws.insert_rows(insert_row)
        prev_seq_value = ws.cell(row=insert_row - 1, column=1).value
        
        if isinstance(prev_seq_value, (int, float)):
            next_no = int(prev_seq_value) + 1
        else:
            # Fallback in case the table is completely empty and has no prior numbers
            next_no = 1
        
        final_datetime_obj = None

        if raw_timestamp:
            try:
                # Clean up duplicate spaces or punctuation traps
                clean_ts = raw_timestamp.strip().replace(" ,", ",").replace(", ", ",")
                
                # Standardize to "Date Time" format
                if "," in clean_ts:
                    time_part, date_part = clean_ts.split(",")
                    combined = f"{date_part.strip()} {time_part.strip()}"
                else:
                    combined = clean_ts

                # List of potential layouts (handles both 12-hour AM/PM and 24-hour formats)
                formats_to_try = [
                    "%d/%m/%Y %I:%M %p",     # Example: 20/5/2026 10:08 AM
                    "%d/%m/%Y %I:%M:%S %p",  # Example: 20/5/2026 10:08:00 AM
                    "%d/%m/%Y %H:%M",        # Example: 6/8/2026 15:42 (Your new format)
                    "%d/%m/%Y %H:%M:%S",     # Example: 6/8/2026 15:42:00
                    "%Y-%m-%d %H:%M:%S"      # Standard server fallback
                ]

                parsed_successfully = False
                for fmt in formats_to_try:
                    try:
                        final_datetime_obj = datetime.strptime(combined, fmt)
                        parsed_successfully = True
                        break # Stop looking once we find a layout that works
                    except ValueError:
                        continue # If it fails, move on to the next format in the list
                
                if not parsed_successfully:
                    raise ValueError(f"Could not match time string: {combined}")

            except Exception as parse_err:
                print(f"[Timestamp Parse Fail]: '{raw_timestamp}' layout mismatch. Fallback to server time. Error: {parse_err}")
                final_datetime_obj = datetime.now()
        else:
            final_datetime_obj = datetime.now()

        # Map metrics cleanly across row coordinates
        ws.cell(row=insert_row, column=1, value=next_no)              # Column A: No
        ws.cell(row=insert_row, column=2, value=data.details)          # Column B: Details
        ws.cell(row=insert_row, column=3, value=data.reported_by)      # Column C: Reported By
        ws.cell(row=insert_row, column=4, value=data.system)           # Column D: System
        ws.cell(row=insert_row, column=5, value=data.category)         # Column E: Category
        ws.cell(row=insert_row, column=6, value="Open")                # Column F: Status
        ws.cell(row=insert_row, column=7, value=data.priority)         # Column G: Priority
        ws.cell(row=insert_row, column=8, value="Kevin")               # Column H: PIC
        # ws.cell(row=insert_row, column=9, value=data.internal_remark)   # Column I: Internal Remark
        
        # Write true date object to Column J and apply custom layout mask string
        date_cell = ws.cell(row=insert_row, column=9, value=final_datetime_obj)
        
        # ---------------------------------------------------------------------
        # NEW FIX: CLONE FORMATTING FROM THE ROW ABOVE
        # ---------------------------------------------------------------------
        # Loop through Columns A to J (1 through 10)
        for col in range(1, 11):
            source_cell = ws.cell(row=insert_row - 1, column=col)
            target_cell = ws.cell(row=insert_row, column=col)

            if source_cell.has_style:
                target_cell.font = copy.copy(source_cell.font)
                target_cell.border = copy.copy(source_cell.border)
                target_cell.fill = copy.copy(source_cell.fill)
                target_cell.number_format = copy.copy(source_cell.number_format)
                target_cell.protection = copy.copy(source_cell.protection)
                target_cell.alignment = copy.copy(source_cell.alignment)
        
        date_cell.number_format = 'd/m/yyyy hh:mm:ss AM/PM'
        
        # ---------------------------------------------------------------------
        # NEW FIX: EXTEND EXCEL TABLE RANGE (BULLETPROOF)
        # ---------------------------------------------------------------------
        # Extract tables safely, whether openpyxl stores them as a dict or list
        all_tables = ws.tables.values() if hasattr(ws.tables, 'values') else ws.tables

        for table in all_tables:
            # Strictly ensure we are working with a Table object, ignoring raw strings
            if hasattr(table, 'ref'):
                min_col, min_row, max_col, max_row = range_boundaries(table.ref)
                
                # If the table ended exactly on the row BEFORE our newly inserted row, stretch it
                if max_row == insert_row - 1:
                    table.ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{insert_row}"

        try:
            wb.save(file_path)
            return next_no
        except PermissionError:
            return "Error: File is currently open in Microsoft Excel. Close it and retry."
            
    except Exception as e:
        return f"Excel Write Error: {str(e)}"       

@app.route('/log-issue', methods=['POST'])
def log_issue():
    try:
        req_data = request.json or {}
        whatsapp_text = req_data.get('text', '')
        browser_sender = req_data.get('chat_sender', 'WhatsApp Client')
        browser_datetime = req_data.get('chat_datetime', '')

        print(f"\n[Incoming Payload] Sender: {browser_sender} | Time: {browser_datetime}")

        api_key = os.environ.get("DEEPSEEK_API_KEY")
        if not api_key:
            return jsonify({"status": "error", "message": "DEEPSEEK_API_KEY variable is missing on host terminal environment session."}), 500

        client = OpenAI(base_url="https://api.deepseek.com", api_key=api_key)

        # Added explicit 'json' instructions to ensure strict DeepSeek API schema compliance
        system_instructions = f"""
        You are an expert technical triage assistant. Your task is to output a strictly formatted json object.
        You are extracting parameters from a user chat problem statement.
        
        Use these contextual parameters extracted directly from the browser DOM:
        - Suggested Sender Identity: {browser_sender}
        - Suggested Message DateTime Context: {browser_datetime}

        CRITICAL STRUCTURAL MAPPING FOR THE JSON OBJECT:
        - 'reported_by': Use the Suggested Sender Identity. Clean it up if it contains mobile numbers or usernames.
        - 'system': Map strictly to 'SMART', 'CFF', or 'eApp & eTax'. Deduce based on contextual feature markers.
        - 'category': Must be strictly 'Issue Support', 'Clarification', or 'Enhancement'.
        - 'priority': Assign 1 for breaking 500 errors or down services, 2 for standard operational bugs.
        - 'details': Synthesize a clean 1-sentence engineering summary describing the error statement.
        """

        response = client.chat.completions.create(
            model="deepseek-v4-flash",
            messages=[
                {"role": "system", "content": system_instructions},
                {"role": "user", "content": f"Parse this message block:\n\n{whatsapp_text}"}
            ],
            response_format={"type": "json_object"},
            temperature=0.1
        )

        raw_ai_output = response.choices[0].message.content
        print(f"[DeepSeek Response Payload]: {raw_ai_output}")

        parsed_json = json.loads(raw_ai_output)
        validated_data = MandayLogExtractor(**parsed_json)
        
        # Name of your master tracking file
        workbook_path = "IssueLog.xlsx"
        ticket_id = append_to_actual_manday_sheet(workbook_path, validated_data, browser_datetime)

        if isinstance(ticket_id, str) and "Error" in ticket_id:
            print(f"[Spreadsheet Execution Failure]: {ticket_id}")
            return jsonify({"status": "error", "message": ticket_id}), 500

        print(f"[Success Tracking Pipeline]: Logged entry row item #{ticket_id}")
        return jsonify({"status": "success", "message": f"Successfully logged Issue #{ticket_id}"}), 200

    except Exception as e:
        print("\n!!! EXCEPTION CAUGHT !!!")
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500

if __name__ == '__main__':
    app.run(port=5000, debug=True)